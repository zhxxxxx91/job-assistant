#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AI求职助手 - Streamlit Web UI"""

import os
import re
import smtplib
import tempfile
import time
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openai
import openpyxl
import PyPDF2
import streamlit as st

# ── 页面配置 ──────────────────────────────────────────────
st.set_page_config(page_title="爽投投 - AI求职助手", page_icon="🎯", layout="wide")

# OpenAI兼容API配置
API_KEY = os.getenv("API_KEY", "")
API_BASE = os.getenv("API_BASE", "https://bobdong.cn/v1")
MODEL_NAME = os.getenv("MODEL_NAME", "Kimi-K2.5")

if not API_KEY:
    st.error("未配置API Key，请联系管理员")
    st.stop()

client = openai.OpenAI(api_key=API_KEY, base_url=API_BASE)

# ══════════════════════════════════════════════════════════
# 工具函数（必须在使用前定义）
# ══════════════════════════════════════════════════════════

@st.cache_data
def extract_user_info_from_resume(pdf_bytes):
    """用AI从简历PDF提取个人信息"""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes)
        tmp_path = f.name
    
    try:
        reader = PyPDF2.PdfReader(tmp_path)
        text = "\n".join([page.extract_text() for page in reader.pages])
    finally:
        os.unlink(tmp_path)
    
    prompt = f"""从以下简历中提取个人基本信息，返回JSON格式。

简历内容：
{text[:2000]}

提取字段：
- name: 姓名
- school: 学校（本科或最高学历）
- major: 专业
- grade: 年级（如"大三"、"研一"，如果是应届生写"应届"）
- grad_year: 毕业时间（格式：YYYY年MM月）
- intern_period: 可实习时间（如果简历中有提及，否则留空）

返回JSON：
{{
  "name": "姓名",
  "school": "学校",
  "major": "专业",
  "grade": "年级",
  "grad_year": "毕业时间",
  "intern_period": "可实习时间或空"
}}"""

    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=300,
        temperature=0.3
    )
    
    import json
    text = response.choices[0].message.content.strip()
    if "```json" in text:
        text = text.split("```json")[1].split("```")[0].strip()
    elif "```" in text:
        text = text.split("```")[1].split("```")[0].strip()
    
    return json.loads(text)


@st.cache_data
def extract_resume_highlights(pdf_bytes):
    """用AI从PDF提取3-5个核心亮点"""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes)
        tmp_path = f.name
    
    try:
        reader = PyPDF2.PdfReader(tmp_path)
        text = "\n".join([page.extract_text() for page in reader.pages])
    finally:
        os.unlink(tmp_path)
    
    prompt = f"""从以下简历中提取3-5个最核心的亮点，每个亮点一句话，用于求职邮件。
要求：
1. 突出量化成果（数字、项目数量、金额等）
2. 突出相关经验（实习、项目、技能）
3. 每个亮点15-25字

简历内容：
{text[:3000]}

请直接输出亮点列表，每行一个，格式：
- 亮点1
- 亮点2
- 亮点3"""

    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=500,
        temperature=0.7
    )
    return response.choices[0].message.content.strip()


@st.cache_data
def load_jobs(excel_bytes):
    """AI自动识别Excel格式并读取岗位"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(excel_bytes)
        tmp_path = f.name
    
    wb = openpyxl.load_workbook(tmp_path)
    ws = wb.active
    
    # 读取前10行用于AI分析格式
    sample_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 10:
            sample_rows.append(row)
        else:
            break
    
    # AI识别格式
    sample_text = "\n".join([f"第{i+1}行: {row}" for i, row in enumerate(sample_rows)])
    
    prompt = f"""分析Excel表格格式，识别哪一列是JD描述，哪一列是邮箱，从第几行开始是数据。

前10行内容：
{sample_text}

返回JSON：
{{
  "jd_column": JD描述所在列号（从0开始，如第3列=2）,
  "email_column": 邮箱所在列号（从0开始）,
  "start_row": 数据起始行号（从1开始）
}}

提示：
- JD列通常包含公司名、职位、职责等长文本
- 邮箱列包含@符号的邮箱地址
- 数据通常从第3-5行开始（前面可能是标题）"""

    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=200,
        temperature=0.3
    )
    
    import json
    text = response.choices[0].message.content.strip()
    if "```json" in text:
        text = text.split("```json")[1].split("```")[0].strip()
    elif "```" in text:
        text = text.split("```")[1].split("```")[0].strip()
    
    format_info = json.loads(text)
    jd_col = format_info["jd_column"]
    email_col = format_info["email_column"]
    start_row = format_info["start_row"]
    
    # 根据识别的格式读取数据
    jobs = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if len(row) <= max(jd_col, email_col):
            continue
        jd_text = row[jd_col]
        email = row[email_col]
        if not jd_text or not email:
            continue
        jobs.append({"jd": str(jd_text).strip(), "email": str(email).strip()})
    
    os.unlink(tmp_path)
    return jobs


def ai_parse_jd(jd_text, name, school, major, grad_year, intern_period):
    """用AI解析JD，提取公司、职位、邮件主题格式、简历命名格式"""
    prompt = f"""解析以下招聘JD，提取关键信息。

JD内容：
{jd_text[:2000]}

请提取：
1. 公司名称（去掉【】等符号）
2. 职位名称
3. 邮件主题格式（如果JD中有明确要求，比如"邮件主题：xxx"，则提取；否则返回null）
4. 简历文件命名格式（如果JD中有明确要求，比如"简历命名：xxx"，则提取；否则返回null）

用户信息：
- 姓名：{name}
- 学校：{school}
- 专业：{major}
- 毕业时间：{grad_year}
- 可实习时间：{intern_period}

返回JSON格式：
{{
  "company": "公司名",
  "position": "职位名",
  "subject_format": "邮件主题格式（如有要求）或null",
  "resume_format": "简历命名格式（如有要求）或null"
}}"""

    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=300,
        temperature=0.7
    )
    
    import json
    text = response.choices[0].message.content.strip()
    if "```json" in text:
        text = text.split("```json")[1].split("```")[0].strip()
    elif "```" in text:
        text = text.split("```")[1].split("```")[0].strip()
    
    result = json.loads(text)
    
    if not result.get("subject_format") or result["subject_format"] == "null":
        result["subject_format"] = f"求职申请-{name}-{school}-{result['position']}"
    else:
        s = result["subject_format"]
        for k, v in {
            "姓名": name, "学校": school, "专业": major,
            "毕业时间": grad_year, "可入职时间": intern_period.split("至")[0] if "至" in intern_period else intern_period,
        }.items():
            s = s.replace(k, v)
        result["subject_format"] = s
    
    if not result.get("resume_format") or result["resume_format"] == "null":
        result["resume_format"] = f"{name}_简历_{result['position']}_{result['company']}.pdf"
    else:
        s = result["resume_format"]
        for k, v in {
            "姓名": name, "学校": school, "专业": major,
            "毕业时间": grad_year, "到岗时间": intern_period.split("至")[0] if "至" in intern_period else intern_period,
        }.items():
            s = s.replace(k, v)
        if not s.endswith(".pdf"):
            s += ".pdf"
        result["resume_format"] = s
    
    return result


def ai_generate_body(company, position, jd_text, resume_highlights, name, school, major, grade, intern_period):
    """用AI生成个性化邮件正文"""
    prompt = f"""为求职邮件生成正文，要求：
1. 100字以内
2. 中文
3. 结构：问候 + 自我介绍 + 1-2个核心匹配点 + 可实习时间 + 期待交流
4. 核心匹配点要结合JD要求和简历亮点

公司：{company}
职位：{position}
JD摘要：{jd_text[:500]}

简历亮点：
{resume_highlights}

个人信息：
- 姓名：{name}
- 学校：{school}
- 专业：{major}
- 年级：{grade}
- 可实习时间：{intern_period}

直接输出邮件正文，不要任何额外说明。"""

    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=300,
        temperature=0.7
    )
    return response.choices[0].message.content.strip()


# ══════════════════════════════════════════════════════════
# 页面布局
# ══════════════════════════════════════════════════════════

st.markdown("<h1 style='font-size: 3rem; margin-bottom: 0;'>🎯 爽投投</h1>", unsafe_allow_html=True)
st.caption("AI智能求职助手 - 精准筛选、智能生成、高效投递 | 原作者：[@Milkyelephants](https://twitter.com/Milkyelephants)")

# ── 侧边栏：用户配置 ──────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 个人配置")
    
    # 简历上传（提前到这里，用于自动填充）
    resume_file_sidebar = st.file_uploader("📄 上传简历 PDF（自动识别信息）", type=["pdf"], key="resume_sidebar")
    
    if resume_file_sidebar and "user_info" not in st.session_state:
        with st.spinner("AI正在识别简历信息..."):
            resume_bytes = resume_file_sidebar.read()
            user_info = extract_user_info_from_resume(resume_bytes)
            st.session_state.user_info = user_info
            st.success("✅ 简历信息已自动识别")
    
    # 使用识别的信息或空值
    default_info = st.session_state.get("user_info", {})
    
    name = st.text_input("姓名", value=default_info.get("name", ""), placeholder="张三")
    school = st.text_input("学校", value=default_info.get("school", ""), placeholder="清华大学")
    major = st.text_input("专业", value=default_info.get("major", ""), placeholder="计算机科学")
    grade = st.text_input("年级", value=default_info.get("grade", ""), placeholder="大三")
    intern_period = st.text_input("可实习时间", value=default_info.get("intern_period", ""), placeholder="2025年7月至10月")
    grad_year = st.text_input("毕业时间", value=default_info.get("grad_year", ""), placeholder="2026年6月")

    st.divider()
    st.header("📧 邮箱配置")
    sender_email = st.text_input("发件邮箱", placeholder="xxxxxx@qq.com")
    auth_code = st.text_input("授权码", type="password", placeholder="QQ邮箱授权码")
    
    with st.expander("❓ 如何获取授权码"):
        st.markdown("""
**QQ邮箱：**
1. 登录 [mail.qq.com](https://mail.qq.com)
2. 设置 → 账户 → POP3/IMAP/SMTP服务
3. 开启服务 → 生成授权码

**163邮箱：**
1. 登录 [mail.163.com](https://mail.163.com)
2. 设置 → POP3/SMTP/IMAP
3. 开启服务 → 新增授权密码

**Gmail：**
1. Google账号 → 安全性
2. 两步验证（必须开启）
3. 应用专用密码 → 生成

**Outlook：**
1. 账户设置 → 安全性
2. 应用密码 → 创建新密码
        """)

    smtp_options = {
        "QQ邮箱 (@qq.com)": ("smtp.qq.com", 465),
        "163邮箱 (@163.com)": ("smtp.163.com", 465),
        "Gmail (@gmail.com)": ("smtp.gmail.com", 465),
        "Outlook (@outlook.com)": ("smtp.office365.com", 587),
    }
    smtp_choice = st.selectbox("邮箱类型", list(smtp_options.keys()))
    smtp_host, smtp_port = smtp_options[smtp_choice]

# ── 主区域：文件上传 ──────────────────────────────────────
st.header("📂 上传岗位Excel")

col1, col2 = st.columns([2, 1])
with col1:
    excel_file = st.file_uploader("📊 岗位Excel", type=["xlsx", "xls"])

with col2:
    with st.expander("📋 Excel格式说明", expanded=False):
        st.markdown("""
**推荐Excel格式：**
- **JD描述列**：包含公司名、职位、职责等详细信息
- **HR邮箱列**：包含@符号的有效邮箱地址
- **数据起始行**：通常从第3-5行开始（前面可能是标题）

**示例格式：**
```
第1行: 标题行
第2行: 说明行  
第3行: 公司A | 产品经理实习生... | hr@companyA.com
第4行: 公司B | 数据分析师... | recruit@companyB.com
```

**AI自动识别：**
系统会自动分析Excel格式，无需固定列位置
        """)

if not resume_file_sidebar:
    st.info("请先在左侧上传简历PDF")
    st.stop()

if not excel_file:
    st.info("请上传岗位Excel")
    st.stop()

# 使用侧边栏上传的简历
resume_bytes = resume_file_sidebar.getvalue()


if not all([name, school, major, grade, intern_period, grad_year]):
    st.warning("请在左侧填写完整个人信息")
    st.stop()

jobs = load_jobs(excel_file.read())
st.success(f"读取到 {len(jobs)} 个岗位")

# 投递模式选择
st.divider()
st.header("📮 选择投递模式")

mode = st.radio(
    "投递方式",
    options=["快速投递", "智能筛选"],
    index=0,
    help="快速投递：直接选择数量投递 | 智能筛选：AI分析分类后精准投递"
)

if mode == "快速投递":
    # 快速模式：直接选数量
    max_rows = st.slider("选择投递岗位数量", 1, len(jobs), min(10, len(jobs)))
    filtered_jobs_data = jobs[:max_rows]
    
    # 转换为统一格式供后续使用
    filtered_jobs = []
    for job in filtered_jobs_data:
        filtered_jobs.append({
            "company": "待解析",
            "position": "待解析", 
            "email": job["email"],
            "jd_full": job["jd"]
        })

else:
    # 智能筛选模式：AI分析分类
    if "parsed_jobs" not in st.session_state:
        with st.spinner("AI正在分析岗位分类..."):
            parsed_jobs = []
            for job in jobs:
                try:
                    parsed = ai_parse_jd(job["jd"], name, school, major, grad_year, intern_period)
                    parsed["email"] = job["email"]
                    parsed["jd_full"] = job["jd"]
                    parsed_jobs.append(parsed)
                except Exception as e:
                    continue
            st.session_state.parsed_jobs = parsed_jobs

    parsed_jobs = st.session_state.parsed_jobs

    # 按行业分类
    categories = {}
    for pj in parsed_jobs:
        company = pj["company"]
        position = pj["position"]
        
        # 根据关键词分类
        if any(k in position.lower() or k in company.lower() for k in ["科技", "tech", "技术", "ai", "算法"]):
            cat = "科技类"
        elif any(k in position.lower() or k in company.lower() for k in ["投资", "pe", "vc", "基金", "资本"]):
            cat = "投资类"
        elif any(k in position.lower() or k in company.lower() for k in ["咨询", "consulting", "战略"]):
            cat = "咨询类"
        elif any(k in position.lower() or k in company.lower() for k in ["金融", "银行", "证券", "保险"]):
            cat = "金融类"
        else:
            cat = "其他"
        
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(pj)

    # 筛选器
    st.subheader("🎯 智能筛选")

    selected_categories = st.multiselect(
        "选择感兴趣的行业类别（可多选）",
        options=list(categories.keys()),
        default=list(categories.keys()),
        help="AI已自动分析岗位类别"
    )

    # 显示每个类别的公司并支持多选
    filtered_jobs = []
    for cat in selected_categories:
        with st.expander(f"📁 {cat} ({len(categories[cat])}个岗位)", expanded=True):
            companies = list(set([pj["company"] for pj in categories[cat]]))
            companies.sort()
            
            selected_companies = st.multiselect(
                f"选择{cat}的公司",
                options=companies,
                default=companies,
                key=f"companies_{cat}",
                help=f"从{len(companies)}家公司中选择"
            )
            
            for pj in categories[cat]:
                if pj["company"] in selected_companies:
                    filtered_jobs.append(pj)

    if not filtered_jobs:
        st.warning("没有符合筛选条件的岗位，请调整筛选条件")
        st.stop()

    st.info(f"筛选后剩余 {len(filtered_jobs)} 个岗位")

    max_rows = st.slider("最多处理岗位数", 1, len(filtered_jobs), min(10, len(filtered_jobs)))
    filtered_jobs = filtered_jobs[:max_rows]

# 将filtered_jobs转回jobs格式供后续使用
jobs = [{"jd": pj["jd_full"], "email": pj["email"]} for pj in filtered_jobs]

# 提取简历亮点（使用已上传的简历）
if "resume_highlights" not in st.session_state:
    with st.spinner("AI正在提取简历亮点..."):
        resume_highlights = extract_resume_highlights(resume_bytes)
        st.session_state.resume_highlights = resume_highlights
else:
    resume_highlights = st.session_state.resume_highlights

st.info(f"简历亮点：\n{resume_highlights}")

# 生成所有岗位数据
if "previews" not in st.session_state:
    with st.spinner("AI正在解析JD并生成邮件..."):
        previews = []
        for job in jobs:
            parsed = ai_parse_jd(job["jd"], name, school, major, grad_year, intern_period)
            body = ai_generate_body(
                parsed["company"], parsed["position"], job["jd"],
                resume_highlights, name, school, major, grade, intern_period
            )
            previews.append({
                "company": parsed["company"],
                "position": parsed["position"],
                "email": job["email"],
                "subject": parsed["subject_format"],
                "resume_name": parsed["resume_format"],
                "body": body,
                "sent": False,
            })
        st.session_state.previews = previews
else:
    previews = st.session_state.previews

# ── 预览表格（可编辑）──────────────────────────────────────
st.divider()
st.subheader("📋 预览与编辑")

for i, p in enumerate(previews):
    with st.expander(f"{'✅' if p['sent'] else '📧'} {i+1}. {p['company']} — {p['position']}", expanded=not p['sent']):
        col1, col2 = st.columns([3, 1])
        with col1:
            st.write(f"**收件人：** `{p['email']}`")
        with col2:
            if not p['sent']:
                if st.button("发送此岗位", key=f"send_{i}", type="primary"):
                    if not sender_email or not auth_code:
                        st.error("请先填写邮箱配置")
                    else:
                        try:
                            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
                                f.write(resume_bytes)
                                tmp_pdf = f.name

                            msg = MIMEMultipart()
                            msg["From"] = sender_email
                            msg["To"] = p["email"]
                            msg["Subject"] = p["subject"]
                            msg.attach(MIMEText(p["body"], "plain", "utf-8"))

                            with open(tmp_pdf, "rb") as f:
                                part = MIMEBase("application", "octet-stream")
                                part.set_payload(f.read())
                            encoders.encode_base64(part)
                            part.add_header("Content-Disposition", "attachment",
                                            filename=("utf-8", "", p["resume_name"]))
                            msg.attach(part)

                            with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
                                server.login(sender_email, auth_code)
                                server.sendmail(sender_email, p["email"], msg.as_string())

                            os.unlink(tmp_pdf)
                            st.session_state.previews[i]["sent"] = True
                            st.success(f"✅ 已发送到 {p['email']}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"发送失败：{e}")
            else:
                st.success("已发送")
        
        # 可编辑字段
        p["subject"] = st.text_input("邮件主题", p["subject"], key=f"subj_{i}")
        p["resume_name"] = st.text_input("附件名称", p["resume_name"], key=f"resume_{i}")
        p["body"] = st.text_area("邮件正文", p["body"], height=200, key=f"body_{i}")

# ── 批量发送区域 ──────────────────────────────────────────
st.divider()
st.subheader("🚀 批量发送")

if not sender_email or not auth_code:
    st.warning("请在左侧填写发件邮箱和授权码")
    st.stop()

unsent = [p for p in previews if not p["sent"]]
if not unsent:
    st.info("所有岗位已发送完毕")
    st.stop()

send_to_self = st.checkbox("📬 先发给自己测试（不发给HR）", value=False)

if st.button(f"批量发送剩余 {len(unsent)} 个岗位", type="primary", use_container_width=True):
    progress = st.progress(0)
    status_box = st.empty()

    for idx, i in enumerate([previews.index(p) for p in unsent]):
        p = previews[i]
        status_box.info(f"正在发送 {idx+1}/{len(unsent)}：{p['company']}...")

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(resume_bytes)
            tmp_pdf = f.name

        to_addr = sender_email if send_to_self else p["email"]

        try:
            msg = MIMEMultipart()
            msg["From"] = sender_email
            msg["To"] = to_addr
            msg["Subject"] = p["subject"]
            msg.attach(MIMEText(p["body"], "plain", "utf-8"))

            with open(tmp_pdf, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment",
                            filename=("utf-8", "", p["resume_name"]))
            msg.attach(part)

            with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
                server.login(sender_email, auth_code)
                server.sendmail(sender_email, to_addr, msg.as_string())

            st.session_state.previews[i]["sent"] = True
            time.sleep(1.5)
        except Exception as e:
            status_box.error(f"❌ {p['company']} 发送失败：{e}")
            time.sleep(2)
        finally:
            os.unlink(tmp_pdf)

        progress.progress((idx + 1) / len(unsent))

    status_box.empty()
    st.success("批量发送完成！")
    st.rerun()
